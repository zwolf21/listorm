import os, io, pathlib
from typing import Text

import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .argtools import select_kwargs



def open_workbook(**kwargs):
    '''create new workbook'''
    return select_kwargs(openpyxl.Workbook, **kwargs)


def load_workbook(file=None, **kwargs):
    if isinstance(file, openpyxl.Workbook):
        return file
    if file is None:
        wb = open_workbook(**kwargs)
        wb.remove(wb.worksheets[0])
        return wb
    if isinstance(file, (str, pathlib.PurePath)):
        if not os.path.exists(file):
            wb = open_workbook(**kwargs)
            wb.remove(wb.worksheets[0])
            return wb
    return select_kwargs(openpyxl.load_workbook, file, **kwargs)


def load_worksheet(workbook:Workbook, sheet_name:Text=None, **kwargs):
    if sheet_name:
        try:
            worksheet = workbook[sheet_name]
        except KeyError:
            worksheet = select_kwargs(workbook.create_sheet, sheet_name, **kwargs)
    else:
        sheets_count = len(workbook.worksheets)
        if sheets_count == 0:
            worksheet = load_worksheet(workbook, 'Sheet1', **kwargs)
        elif sheets_count == 1:
            worksheet = workbook.worksheets[0]
        else:
            worksheet = workbook.active
    return worksheet


def save_workbook(workbook:Workbook, file=None, close=True):
    f = file or io.BytesIO()

    if isinstance(f, str):
        dirname = os.path.dirname(f)
        if dirname:
            if not os.path.exists(dirname):
                os.makedirs(dirname, exist_ok=True)

    workbook.save(f)
    if close:
        workbook.close()
    if file is None and isinstance(f, io.BufferedIOBase):
        return f.getvalue()
    return file


def set_cell_height(worksheet:Worksheet, rows:int, height:float):
    worksheet.row_dimensions[rows].auto_size = False
    worksheet.row_dimensions[rows].height = height


def set_cell_width(worksheet:Worksheet, cols:int, width:float):
    worksheet.column_dimensions[get_column_letter(cols)].auto_size = False
    worksheet.column_dimensions[get_column_letter(cols)].width = width


def add_image(worksheet:Worksheet, content:bytes, rows:int, cols:int, size:tuple=None):
    fp = io.BytesIO(content)
    cell_address = f'{get_column_letter(cols)}{rows}'
    image = Image(fp)
    if size and len(size) == 2:
        image.height, image.width = size
    worksheet.add_image(image, cell_address)


    

